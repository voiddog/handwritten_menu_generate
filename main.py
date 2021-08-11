import random

from PIL import Image, ImageFont, ImageColor
from openpyxl import Workbook, load_workbook
from DrawBox import *
from FoodMenu import *

Offset = tuple[float, float]


# 菜单单元生成配置
class FoodMenuGenConfig:

    def __init__(self, offset: Offset, kind: str, count: int):
        super().__init__()
        self.kind = kind
        self.count = count
        self.offset = offset


# 将 excel 菜单列表解析成数据列表
# 格式：菜名 - 最低价 - 最高价 - 类型
def _parseFoodList(excel_file_path) -> list[FoodItem]:
    work_book = load_workbook(filename=excel_file_path)
    sheet = work_book.active
    ret = []
    for row in sheet.iter_rows(min_row=2):
        ret.append(FoodItem(
            name=row[0].value,
            min_price=row[1].value,
            max_price=row[2].value,
            kind=row[3].value
        ))
    return ret

# 生成随机的单个菜单生成配置
# start_offset 和 end_offset 决定了绘制区间左上角的位置随机区域
# min_count 和 max_count 决定菜单中菜的个数


def _generateSingleMenuGenConfig(
    start_offset: Offset,
    end_offset: Offset,
    kind: str,
    min_count: int,
    max_count: int,
) -> FoodMenuGenConfig:
    return FoodMenuGenConfig(
        (random.randint(start_offset[0], end_offset[0]),
         random.randint(start_offset[1], end_offset[1])),
        kind,
        random.randint(min_count, max_count)
    )


def _generateDayOrderGenerator() -> list[FoodMenuGenConfig]:
    ret = []
    # 肉类
    ret.append(_generateSingleMenuGenConfig((20, 40), (40, 80), "肉类", 2, 4))
    # 水产类
    ret.append(_generateSingleMenuGenConfig((30, 400), (50, 450), "水产", 5, 7))
    # 蔬菜类
    ret.append(_generateSingleMenuGenConfig((420, 40), (450, 60), "蔬菜", 9, 12))
    return ret


def _generateDrawBoxWithFont(font_name: str) -> DrawBox:
    return DrawBox(
        font=ImageFont.truetype(font_name, size=60),
        fill=ImageColor.getrgb("#333"),
        line_spacing=80,
        word_spacing=0,
        line_spacing_sigma=5,  # 行间距随机扰动
        font_size_sigma=5,  # 字体大小随机扰动
        perturb_x_sigma=0,  # 笔画横向偏移随机扰动
        perturb_y_sigma=0,  # 笔画纵向偏移随机扰动
        perturb_theta_sigma=0.1,  # 笔画旋转偏移随机扰动
    )


def _renderDayOrder(
        draw_box: DrawBox,
        page_size: Offset,
        config_list: list[FoodMenuGenConfig],
        price_offset: Offset,
        food_item_map: dict,
) -> tuple[Image, int]:
    page = Image.new(mode="RGBA", size=page_size,
                     color=ImageColor.getrgb("#ffffff00"))
    price_sum = 0
    for config in config_list:
        (cell_img, cell_price_sum) = renderRandomMenuList(
            draw_box,
            food_item_map[config.kind],
            config.count,
        )
        page.paste(cell_img, config.offset, cell_img)
        price_sum += cell_price_sum

    # 添加总价格
    price_sum_img = draw_box.render("总计 %s" % price_sum)
    page.paste(price_sum_img, price_offset, price_sum_img)
    return (page, price_sum)


if __name__ == '__main__':
    # parse food menu
    food_item_list = _parseFoodList('menu.xlsx')
    # splite list to map
    food_item_map = {}
    for food_item in food_item_list:
        kind = food_item.kind

        if kind not in food_item_map:
            food_item_map[kind] = []

        food_item_map[kind].append(food_item)

    # make draw box list
    draw_box_list = []
    for i in range(1, 6):
        draw_box_list.append(
            _generateDrawBoxWithFont("shouxie-%s.ttf" % i)
        )

    count = 20
    for i in range(0, count):
        print("处理第 %s 张" % i)
        draw_index = random.randint(0, len(draw_box_list) - 1)
        draw_box = draw_box_list[draw_index]
        # 第一天
        day_1 = _renderDayOrder(
            draw_box,
            (877, 1240),
            _generateDayOrderGenerator(),
            (random.randint(300, 350), random.randint(1000, 1100)),
            food_item_map)[0]
        # 第二天
        day_2 = _renderDayOrder(
            draw_box,
            (877, 1240),
            _generateDayOrderGenerator(),
            (random.randint(300, 350), random.randint(1000, 1100)),
            food_item_map)[0]
        # merge
        page = Image.new(mode="RGBA", size=(1754, 1240),
                         color=ImageColor.getrgb("#FFFFFF"))
        page.paste(day_1, (0, 0), day_1)
        page.paste(day_2, (877, 0), day_2)
        page.save("out/%s.png" % i, format="PNG")
